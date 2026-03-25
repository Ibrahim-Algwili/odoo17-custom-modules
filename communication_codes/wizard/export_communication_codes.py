# -*- coding: utf-8 -*-
import html2text

from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
from datetime import datetime


class ExportCommunicationCodesWizard(models.TransientModel):
    _name = 'export.communication.codes.wizard'
    _description = 'تصدير شفرات الاتصال إلى Excel'
    
    code_ids = fields.Many2many(
        'communication.codes',
        string='الشفرات',
        default=lambda self: self._default_code_ids(),
    )
    
    export_all = fields.Boolean(
        string='تصدير جميع السجلات',
        default=True,
    )
    
    file = fields.Binary(
        string='ملف Excel',
        readonly=True,
    )
    
    filename = fields.Char(
        string='اسم الملف',
        readonly=True,
    )
    
    state = fields.Selection(
        [
            ('draft', 'مسودة'),
            ('done', 'تم'),
        ],
        string='الحالة',
        default='draft',
        readonly=True,
    )
    
    def _default_code_ids(self):
        if self._context.get('active_ids'):
            return self._context.get('active_ids')
        return []
    
    def action_export(self):
        self.ensure_one()
        
        codes = self.code_ids
        if self.export_all:
            codes = self.env['communication.codes'].search([])
        
        if not codes:
            raise UserError('لا توجد بيانات للتصدير!')
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise UserError('مكتبة openpyxl غير مثبتة! يرجى تثبيتها.')
        
        # إنشاء ملف Excel
        wb = Workbook()
        ws = wb.active
        ws.title = 'شفرات الاتصالات'
        
        # تنسيق الرأس
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        
        # رؤوس الأعمدة
        headers = [
            'الرقم التسلسلي',
            'اسم الموظف',
            'الوظيفة',
            'الشركة',
            'الفرع',
            'المدينة',
            'رقم الشفرة',
            'نظام الشفرة',
            'الرصيد الشهري',
            'حالة الشفرة',
            'إصدار الشفرة',
            'تاريخ التسليم',
            'ملاحظات',
        ]


        # تحويل القيم للعرض
        code_system_map = {
            'prepaid': 'دفع مسبق',
            'monthly_invoice': 'فاتورة شهرية',
            'other': 'أخرى',
        }

        code_status_map = {
            'in_stock': 'في المخزن',
            'delivered': 'تم التسليم',
            'suspended': 'موقوفة',
            'cancelled': 'ملغاة',
        }
        
        code_version_map = {
            'original': 'الأصلي',
            'new_version': 'إصدار شفرة جديد',
        }

        # إضافة رؤوس الأعمدة
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # تعريف تنسيق البيانات (أضفه مع تعريفات الـ Header في الأعلى)
        data_alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

        # إضافة البيانات داخل الـ Loop
        for row_idx, code in enumerate(codes, start=2):
            # مصفوفة القيم لتسهيل التكرار والتنسيق
            row_values = [
                code.name or '',
                code.employee_id.name or '',
                code.job_id.name or '',
                code.company_id.name or '',
                code.branch_id.name or '',
                code.city or '',
                code.code_number or '',
                code_system_map.get(code.code_system, ''),
                code.monthly_balance or 0,
                code_status_map.get(code.code_status, ''),
                code_version_map.get(code.code_version, ''),
                str(code.delivery_date) if code.delivery_date else '',
                # تنظيف النص من المسافات والأسطر الزائدة الناتجة عن HTML
                html2text.html2text(code.notes or '').strip()
            ]

            for col_idx, value in enumerate(row_values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # تطبيق التوسيط على كل خلية في السطر
                cell.alignment = data_alignment

            #  زيادة ارتفاع السطر ليكون التوسيط العمودي واضحاً
            ws.row_dimensions[row_idx].height = 35

        
        # تعديل عرض الأعمدة
        column_widths = [15, 20, 20, 20, 20, 15, 20, 20, 15, 15, 20, 20, 30]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[chr(64 + col_idx)].width = width
        
        # حفظ الملف
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # تحويل إلى base64
        file_data = base64.b64encode(output.getvalue())
        
        # إنشاء اسم الملف
        filename = f'شفرات_الاتصالات_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        self.write({
            'file': file_data,
            'filename': filename,
            'state': 'done',
        })
        
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'export.communication.codes.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
    
    def action_download(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{self._name}/{self.id}/file/{self.filename}?download=true',
            'target': 'self',
        }
    
    def action_reset(self):
        self.write({
            'state': 'draft',
            'file': False,
            'filename': False,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'export.communication.codes.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

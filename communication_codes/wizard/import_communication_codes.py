# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import ValidationError, UserError
import base64
import io
from datetime import datetime


class ImportCommunicationCodesWizard(models.TransientModel):
    _name = 'import.communication.codes.wizard'
    _description = 'استيراد شفرات الاتصال من Excel'
    
    file = fields.Binary(
        string='ملف Excel',
        required=True,
        attachment=False,
    )
    
    filename = fields.Char(
        string='اسم الملف',
    )
    
    import_mode = fields.Selection(
        [
            ('create', 'إنشاء سجلات جديدة'),
            ('update', 'تحديث السجلات الموجودة'),
        ],
        string='وضع الاستيراد',
        default='create',
        required=True,
    )
    
    error_lines = fields.Text(
        string='سجل الأخطاء',
        readonly=True,
    )
    
    success_count = fields.Integer(
        string='عدد السجلات الناجحة',
        readonly=True,
        default=0,
    )
    
    error_count = fields.Integer(
        string='عدد السجلات الفاشلة',
        readonly=True,
        default=0,
    )
    
    state = fields.Selection(
        [
            ('draft', 'مسودة'),
            ('done', 'تم'),
        ],
        string='الحالة',
        default='draft',
        readonly=True,
    )
    
    def action_import(self):
        self.ensure_one()
        
        if not self.file:
            raise UserError('الرجاء اختيار ملف Excel!')
        
        if not self.filename or not self.filename.endswith('.xlsx'):
            raise UserError('الرجاء اختيار ملف Excel صالح (.xlsx)!')
        
        try:
            # قراءة الملف
            file_content = base64.b64decode(self.file)
            workbook = self.import_excel(file_content)
            
            # التحقق من صحة البيانات
            errors = []
            success_count = 0
            error_count = 0
            
            # الحصول على Sheet الأول
            sheet = workbook.active
            
            # التحقق من رأس الجدول
            headers = [cell.value for cell in sheet[1]]
            required_columns = [
                'اسم الموظف', 'الوظيفة', 'الشركة', 'الفرع', 
                'المدينة', 'رقم الشفرة', 'نظام الشفرة', 
                'الرصيد الشهري', 'حالة الشفرة', 'إصدار الشفرة'
            ]
            
            # التحقق من وجود الأعمدة المطلوبة
            for col in required_columns:
                if col not in headers:
                    errors.append(f'العمود "{col}" غير موجود في الملف!')
            
            if errors:
                self.write({
                    'error_lines': '\n'.join(errors),
                    'state': 'done',
                    'error_count': len(errors),
                })
                return {
                    'type': 'ir.actions.act_window',
                    'res_model': 'import.communication.codes.wizard',
                    'res_id': self.id,
                    'view_mode': 'form',
                    'target': 'new',
                }
            
            # قراءة البيانات
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0] and not row[5]:  # Skip empty rows
                    continue
                
                try:
                    # استخراج البيانات
                    employee_name = row[0] if row[0] else ''
                    # الأعمدة التالية موجودة في الملف لأغراض توثيقية ولكن تعتمد قيمها في السجل
                    # على بيانات الموظف نفسه (حقول related)، لذلك لا تُخزن مباشرة في الموديل.
                    job_title = row[1] if row[1] else ''
                    company_name = row[2] if row[2] else ''
                    branch_name = row[3] if row[3] else ''
                    city = row[4] if row[4] else ''
                    code_number = str(row[5]).strip() if row[5] else ''
                    code_system = row[6] if row[6] else 'prepaid'
                    monthly_balance = row[7]
                    code_status = row[8] if row[8] else 'in_stock'
                    code_version = row[9] if row[9] else 'original'
                    
                    # التحقق من رقم الشفرة
                    if not code_number:
                        errors.append(f'السطر {row_idx}: رقم الشفرة مطلوب!')
                        error_count += 1
                        continue
                    
                    # التحقق من عدم تكرار رقم الشفرة
                    existing_code = self.env['communication.codes'].search([
                        ('code_number', '=', code_number)
                    ], limit=1)
                    
                    if existing_code and self.import_mode == 'create':
                        errors.append(f'السطر {row_idx}: رقم الشفرة "{code_number}" مستخدم بالفعل!')
                        error_count += 1
                        continue
                    
                    # البحث عن الموظف
                    employee_id = False
                    if employee_name:
                        employee = self.env['hr.employee'].search([
                            ('name', 'ilike', employee_name)
                        ], limit=1)
                        if employee:
                            employee_id = employee.id
                        else:
                            # إنشاء موظف جديد إذا لم يوجد
                            employee_id = self.env['hr.employee'].create({
                                'name': employee_name,
                            }).id
                    
                    if not employee_id:
                        errors.append(f'السطر {row_idx}: الموظف "{employee_name}" غير موجود!')
                        error_count += 1
                        continue
                    
                    # تحويل نظام الشفرة
                    code_system_map = {
                        'دفع مسبق': 'prepaid',
                        'فاتورة شهرية': 'monthly_invoice',
                        'أخرى': 'other',
                        'prepaid': 'prepaid',
                        'monthly_invoice': 'monthly_invoice',
                        'other': 'other',
                    }
                    if code_system not in code_system_map:
                        errors.append(f'السطر {row_idx}: قيمة نظام الشفرة "{code_system}" غير صحيحة!')
                        error_count += 1
                        continue
                    code_system_val = code_system_map[code_system]
                    
                    # تحويل حالة الشفرة
                    code_status_map = {
                        'في المخزن': 'in_stock',
                        'تم التسليم': 'delivered',
                        'موقوفة': 'suspended',
                        'ملغاة': 'cancelled',
                        'in_stock': 'in_stock',
                        'delivered': 'delivered',
                        'suspended': 'suspended',
                        'cancelled': 'cancelled',
                    }
                    if code_status not in code_status_map:
                        errors.append(f'السطر {row_idx}: قيمة حالة الشفرة "{code_status}" غير صحيحة!')
                        error_count += 1
                        continue
                    code_status_val = code_status_map[code_status]
                    
                    # تحويل إصدار الشفرة
                    code_version_map = {
                        'الأصلي': 'original',
                        'إصدار شفرة جديد': 'new_version',
                        'original': 'original',
                        'new_version': 'new_version',
                    }
                    if code_version not in code_version_map:
                        errors.append(f'السطر {row_idx}: قيمة إصدار الشفرة "{code_version}" غير صحيحة!')
                        error_count += 1
                        continue
                    code_version_val = code_version_map[code_version]
                    
                    # إنشاء أو تحديث السجل
                    vals = {
                        'employee_id': employee_id,
                        'city': city or '',
                        'code_number': code_number,
                        'code_system': code_system_val,
                        'monthly_balance': float(monthly_balance) if monthly_balance else 0.0,
                        'code_status': code_status_val,
                        'code_version': code_version_val,
                    }
                    
                    if existing_code:
                        existing_code.write(vals)
                    else:
                        self.env['communication.codes'].create(vals)
                    
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f'السطر {row_idx}: خطأ في البيانات - {str(e)}')
                    error_count += 1
            
            # حفظ النتيجة
            error_message = '\n'.join(errors) if errors else 'تم الاستيراد بنجاح!'
            
            self.write({
                'error_lines': error_message,
                'success_count': success_count,
                'error_count': error_count,
                'state': 'done',
            })
            
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'import.communication.codes.wizard',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
            }
            
        except Exception as e:
            raise UserError(f'خطأ في قراءة الملف: {str(e)}')
    
    def import_excel(self, file_content):
        """استيراد ملف Excel باستخدام openpyxl"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise UserError('مكتبة openpyxl غير مثبتة! يرجى تثبيتها.')
        
        return load_workbook(io.BytesIO(file_content))
    
    def action_reset(self):
        self.write({
            'state': 'draft',
            'file': False,
            'filename': False,
            'error_lines': False,
            'success_count': 0,
            'error_count': 0,
        })
        return {
            'type': 'ir.actions.act_window',
            'res_model': 'import.communication.codes.wizard',
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
